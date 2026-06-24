from __future__ import annotations

import re
from datetime import date, datetime
from pathlib import Path

import openpyxl
import pandas as pd

from viewership_model.data.arizona_import import _coerce_date, _parse_season, _parse_week
from viewership_model.data.d1_teams import normalize_team_name
from viewership_model.data.research_import import classify_game_type

ESPN_PLUS = "ESPN+"

DEAL_SPORT_MAP: dict[str, str] = {
    "ncaa men's basketball": "mens_basketball",
    "ncaa women's basketball": "womens_basketball",
    "ncaa baseball": "baseball",
    "ncaa softball": "softball",
    "ncaa football": "football",
}

HEADER_ALIASES: dict[str, tuple[str, ...]] = {
    "deal": ("DEAL",),
    "title": ("AIRING_TITLE", "TITLE", "PROGRAM_TITLE"),
    "date": ("ORIGINAL_AIR_DATE_EST",),
    "conference": ("HOME_TEAM_CONFERENCE", "CONFERENCE"),
    "away_team": ("AWAY_TEAM",),
    "home_team": ("HOME_TEAM",),
    "unique_viewers": ("UNIQUE_VIEWERS",),
    "ratings_id": ("RATINGS_ID", "Ratings_id"),
}


def _header_map(ws) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        value = ws.cell(1, col).value
        if value is None:
            continue
        text = str(value).strip()
        mapping[text] = col
        mapping[text.upper()] = col
    return mapping


def _header_col(headers: dict[str, int], *names: str) -> int | None:
    for name in names:
        if name in headers:
            return headers[name]
        upper = name.upper()
        if upper in headers:
            return headers[upper]
    lowered = {key.strip().lower(): idx for key, idx in headers.items()}
    for name in names:
        idx = lowered.get(name.strip().lower())
        if idx is not None:
            return idx
    return None


def _resolve_columns(headers: dict[str, int]) -> dict[str, int | None]:
    return {
        key: _header_col(headers, *aliases)
        for key, aliases in HEADER_ALIASES.items()
    }


def _clean_team_name(value: object) -> str | None:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    text = str(value).strip()
    if not text:
        return None
    text = re.sub(r"^#\d+\s+", "", text)
    return normalize_team_name(text)


def _sport_from_deal(deal: object) -> str | None:
    if deal is None or (isinstance(deal, float) and pd.isna(deal)):
        return None
    return DEAL_SPORT_MAP.get(str(deal).strip().lower())


def _team_involved(team: str | None, school: str) -> bool:
    if not team:
        return False
    return team.strip().lower() == school.strip().lower()


def _location_type(home_team: str, away_team: str, school: str) -> str:
    if home_team.strip().lower() == school.strip().lower():
        return "home"
    if away_team.strip().lower() == school.strip().lower():
        return "away"
    return "neutral"


def _numeric_viewers(value: object) -> float | None:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    try:
        viewers = float(value)
    except (TypeError, ValueError):
        return None
    if viewers <= 0:
        return None
    return viewers


def _read_sheet(
    ws,
    source_sheet: str,
    *,
    home_team: str,
    game_id_prefix: str,
    default_conference: str,
    home_city: str,
    home_state: str,
    sports: set[str] | None,
    workbook_name: str,
    row_offset: int,
) -> tuple[list[dict], int]:
    headers = _header_map(ws)
    cols = _resolve_columns(headers)
    if None in (
        cols["deal"],
        cols["away_team"],
        cols["home_team"],
        cols["unique_viewers"],
        cols["date"],
    ):
        return [], row_offset

    school = normalize_team_name(home_team)
    rows: list[dict] = []
    next_offset = row_offset

    for row_idx in range(2, ws.max_row + 1):
        sport = _sport_from_deal(ws.cell(row_idx, cols["deal"]).value)
        if sport is None:
            continue
        if sports and sport not in sports:
            continue

        away_team = _clean_team_name(ws.cell(row_idx, cols["away_team"]).value)
        row_home_team = _clean_team_name(ws.cell(row_idx, cols["home_team"]).value)
        viewers = _numeric_viewers(ws.cell(row_idx, cols["unique_viewers"]).value)
        if away_team is None or row_home_team is None or viewers is None:
            continue
        if not (_team_involved(away_team, school) or _team_involved(row_home_team, school)):
            continue

        game_date_raw = ws.cell(row_idx, cols["date"]).value
        game_date = _coerce_date(game_date_raw)
        season = _parse_season(game_date_raw)
        week = _parse_week(game_date_raw)
        location_type = _location_type(row_home_team, away_team, school)

        title_col = cols["title"]
        title = ""
        if title_col is not None:
            title_value = ws.cell(row_idx, title_col).value
            title = str(title_value).strip() if title_value is not None else ""

        conf_col = cols["conference"]
        conference = default_conference
        if conf_col is not None:
            conf_value = ws.cell(row_idx, conf_col).value
            if conf_value not in (None, ""):
                conference = str(conf_value).strip()
                if "atlantic 10" in conference.lower():
                    conference = default_conference

        ratings_col = cols["ratings_id"]
        ratings_id = ""
        if ratings_col is not None:
            ratings_value = ws.cell(row_idx, ratings_col).value
            ratings_id = str(ratings_value).strip() if ratings_value is not None else ""

        game_type = classify_game_type(
            sport=sport,
            notes=title,
            week=week,
            network=ESPN_PLUS,
            viewership_millions=viewers / 1_000_000,
        )
        opponent = away_team if row_home_team == school else row_home_team

        next_offset += 1
        rows.append(
            {
                "game_id": f"{game_id_prefix}-{sport}-{season}-{next_offset:04d}",
                "sport": sport,
                "season": season,
                "week": week,
                "home_team": row_home_team,
                "away_team": away_team,
                "network": ESPN_PLUS,
                "conference": conference,
                "location_type": location_type,
                "location_city": home_city if location_type == "home" else "Unknown",
                "location_state": home_state if location_type == "home" else "ST",
                "is_rivalry": 0,
                "is_ranked_matchup": 0,
                "is_prime_time": 0,
                "viewership_millions": round(viewers / 1_000_000, 6),
                "avg_viewers": viewers,
                "is_estimate": 0,
                "game_date": game_date,
                "gender": "",
                "opponent": opponent,
                "source_sheet": f"{workbook_name}:{source_sheet}",
                "game_type": game_type,
                "ratings_id": ratings_id,
            }
        )
    return rows, next_offset


def import_espn_dtc_workbook(
    workbook_path: Path | str,
    *,
    home_team: str,
    sheets: list[str] | None = None,
    game_id_prefix: str = "SB",
    default_conference: str = "Atlantic 10",
    home_city: str = "St. Bonaventure",
    home_state: str = "NY",
    sports: list[str] | None = None,
) -> pd.DataFrame:
    path = Path(workbook_path)
    wb = openpyxl.load_workbook(path, data_only=True, read_only=False)
    sheet_names = sheets or wb.sheetnames
    sport_filter = set(sports) if sports else None

    all_rows: list[dict] = []
    row_offset = 0
    for sheet_name in sheet_names:
        if sheet_name not in wb.sheetnames:
            continue
        sheet_rows, row_offset = _read_sheet(
            wb[sheet_name],
            sheet_name,
            home_team=home_team,
            game_id_prefix=game_id_prefix,
            default_conference=default_conference,
            home_city=home_city,
            home_state=home_state,
            sports=sport_filter,
            workbook_name=path.name,
            row_offset=row_offset,
        )
        all_rows.extend(sheet_rows)

    wb.close()
    if not all_rows:
        return pd.DataFrame()

    df = pd.DataFrame(all_rows)
    return df.drop_duplicates(
        subset=["sport", "game_date", "home_team", "away_team", "network", "location_type"],
        keep="first",
    ).reset_index(drop=True)
