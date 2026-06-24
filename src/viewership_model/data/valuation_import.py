from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime, time
from pathlib import Path

import openpyxl
import pandas as pd

from viewership_model.data.arizona_import import (
    SPORT_ALIASES,
    YELLOW_RGB,
    _cell_is_yellow,
    _coerce_date,
    _header_col,
    _header_map,
    _is_prime_time,
    _location_type,
    _normalize_sport,
    _numeric_value,
    _parse_season,
    _parse_week,
)
from viewership_model.data.d1_teams import build_d1_teams, normalize_team_name
from viewership_model.data.key_tab_import import (
    build_networks_from_key,
    canonicalize_network_name,
    merge_key_tabs,
)
from viewership_model.data.espn_dtc_import import import_espn_dtc_workbook
from viewership_model.data.monmouth_import import import_monmouth_schedule_workbook
from viewership_model.data.network_reach import compute_network_reach_scores


def _normalize_network(value: str | None) -> str | None:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    text = str(value).strip()
    if not text:
        return None
    return canonicalize_network_name(text)

ARIZONA_PRIMARY_SHEET = "Valuation (24-25 Data)"
NJIT_PRIMARY_SHEET = "NJIT Jersey Patch Valuation"


@dataclass
class ValuationImportConfig:
    home_team: str
    sheets: list[str]
    game_id_prefix: str
    default_conference: str
    home_city: str
    home_state: str
    format: str = "standard"


DEFAULT_VALUATION_WORKBOOKS: list[dict] = [
    {
        "file": "data/Arizona Jersey Patch Valuation .xlsx",
        "home_team": "Arizona",
        "sheets": [ARIZONA_PRIMARY_SHEET],
        "prefix": "AZ",
        "conference": "Big 12",
        "city": "Tucson",
        "state": "AZ",
    },
    {
        "file": "data/NJIT Jersey Patch Valuation 2024-25.xlsx",
        "home_team": "NJIT",
        "sheets": [NJIT_PRIMARY_SHEET],
        "prefix": "NJ",
        "conference": "America East",
        "city": "Newark",
        "state": "NJ",
    },
    {
        "file": "data/Monmouth Jersey Patch Valuation.xlsx",
        "home_team": "Monmouth",
        "format": "monmouth_schedule",
        "sheets": [
            "MBB Schedule 2024-2025",
            "MBB Schedule 2025-26",
            "WBB Schedule 2024-25",
            "WBB Schedule 2025-26",
            "Baseball Schedule 2024-2025",
            "Football Schedule 2024",
            "Football Schedule 2025",
        ],
        "prefix": "MU",
        "conference": "CAA",
        "city": "West Long Branch",
        "state": "NJ",
    },
]


def config_from_dict(entry: dict) -> ValuationImportConfig:
    return ValuationImportConfig(
        home_team=str(entry["home_team"]),
        sheets=list(entry.get("sheets", [])),
        game_id_prefix=str(entry.get("prefix", entry["home_team"][:2].upper())),
        default_conference=str(entry.get("conference", "Unknown")),
        home_city=str(entry.get("city", "Unknown")),
        home_state=str(entry.get("state", "ST")),
        format=str(entry.get("format", "standard")),
    )


def _teams_from_row(opponent: str, home_or_away: str | None, home_team: str) -> tuple[str, str]:
    opponent = str(opponent).strip()
    location = _location_type(home_or_away)
    if location == "away":
        return opponent, home_team
    if location == "neutral":
        return home_team, opponent
    return home_team, opponent


def _read_sheet(
    ws,
    ws_values,
    source_sheet: str,
    config: ValuationImportConfig,
    workbook_name: str,
) -> list[dict]:
    headers = _header_map(ws)
    sport_col = _header_col(headers, "Sport")
    opponent_col = _header_col(headers, "Away ", "Away", "Opponent")
    network_col = _header_col(headers, "Network")
    viewers_col = _header_col(headers, "Avg Viewers")
    if None in (sport_col, opponent_col, network_col, viewers_col):
        missing = []
        if sport_col is None:
            missing.append("Sport")
        if opponent_col is None:
            missing.append("Away/Opponent")
        if network_col is None:
            missing.append("Network")
        if viewers_col is None:
            missing.append("Avg Viewers")
        raise ValueError(f"Sheet '{source_sheet}' missing columns: {missing}")

    gender_col = _header_col(headers, "Gender")
    home_away_col = _header_col(headers, "Home or Away")
    date_col = _header_col(headers, "Date")
    time_col = _header_col(headers, "Time")
    conf_col = _header_col(headers, "Conf")
    conf_type_col = _header_col(headers, "Conference/Non Conference")
    home_rank_col = _header_col(
        headers,
        "Arizona Ranking",
        "NJIT Ranking",
        f"{config.home_team} Ranking",
    )
    away_rank_col = _header_col(headers, "Away Rank", "Opponent Rank")

    rows: list[dict] = []
    for row_idx in range(2, ws.max_row + 1):
        sport_raw = ws.cell(row_idx, sport_col).value
        if sport_raw is None or str(sport_raw).strip() == "":
            continue

        opponent = ws_values.cell(row_idx, opponent_col).value
        network = _normalize_network(ws_values.cell(row_idx, network_col).value)
        viewers_cell = ws.cell(row_idx, viewers_col)
        viewers = _numeric_value(ws_values.cell(row_idx, viewers_col).value)
        if opponent is None or network is None or viewers is None:
            continue

        gender = ws_values.cell(row_idx, gender_col).value if gender_col else None
        sport = _normalize_sport(gender, sport_raw)
        home_or_away = ws_values.cell(row_idx, home_away_col).value if home_away_col else "Home"
        game_date = ws_values.cell(row_idx, date_col).value if date_col else None
        game_time = ws_values.cell(row_idx, time_col).value if time_col else None
        conf = ws_values.cell(row_idx, conf_col).value if conf_col else None
        conf_type = ws_values.cell(row_idx, conf_type_col).value if conf_type_col else None
        home_rank = ws_values.cell(row_idx, home_rank_col).value if home_rank_col else None
        away_rank = ws_values.cell(row_idx, away_rank_col).value if away_rank_col else None

        home_team, away_team = _teams_from_row(str(opponent), home_or_away, config.home_team)
        location_type = _location_type(home_or_away)
        conference = str(conf).strip() if conf not in (None, "") else config.default_conference
        if conf_type and str(conf_type).strip().lower().startswith("non"):
            conference = "Non-Conference"

        season = _parse_season(game_date)
        week = _parse_week(game_date)
        is_ranked = int(
            pd.notna(home_rank) and str(home_rank).strip() != ""
            or pd.notna(away_rank) and str(away_rank).strip() != ""
        )

        rows.append(
            {
                "game_id": f"{config.game_id_prefix}-{sport}-{season}-{row_idx:04d}",
                "sport": sport,
                "season": season,
                "week": week,
                "home_team": home_team,
                "away_team": away_team,
                "network": network,
                "conference": conference,
                "location_type": location_type,
                "location_city": config.home_city if home_team == config.home_team else "Unknown",
                "location_state": config.home_state if home_team == config.home_team else "ST",
                "is_rivalry": 0,
                "is_ranked_matchup": is_ranked,
                "is_prime_time": int(_is_prime_time(game_time)),
                "viewership_millions": round(viewers / 1_000_000, 6),
                "avg_viewers": viewers,
                "is_estimate": int(_cell_is_yellow(viewers_cell)),
                "game_date": _coerce_date(game_date),
                "gender": str(gender or "").strip(),
                "opponent": str(opponent).strip(),
                "source_sheet": f"{workbook_name}:{source_sheet}",
            }
        )
    return rows


def import_valuation_workbook(
    workbook_path: Path | str,
    config: ValuationImportConfig,
    sports: list[str] | None = None,
) -> pd.DataFrame:
    path = Path(workbook_path)
    if config.format == "monmouth_schedule":
        return import_monmouth_schedule_workbook(
            path,
            home_team=config.home_team,
            sheets=config.sheets or None,
            game_id_prefix=config.game_id_prefix,
            default_conference=config.default_conference,
            home_city=config.home_city,
            home_state=config.home_state,
        )

    if config.format == "espn_dtc":
        return import_espn_dtc_workbook(
            path,
            home_team=config.home_team,
            sheets=config.sheets or None,
            game_id_prefix=config.game_id_prefix,
            default_conference=config.default_conference,
            home_city=config.home_city,
            home_state=config.home_state,
            sports=sports,
        )

    wb = openpyxl.load_workbook(path, data_only=False)
    wb_values = openpyxl.load_workbook(path, data_only=True)

    all_rows: list[dict] = []
    for sheet in config.sheets:
        if sheet not in wb.sheetnames:
            continue
        all_rows.extend(
            _read_sheet(wb[sheet], wb_values[sheet], sheet, config, path.name)
        )

    df = pd.DataFrame(all_rows)
    if df.empty:
        return df

    return df.drop_duplicates(
        subset=["sport", "game_date", "home_team", "away_team", "network", "location_type"],
        keep="first",
    ).reset_index(drop=True)


def import_valuation_workbooks(
    workbook_entries: list[tuple[Path | str, ValuationImportConfig]],
    sports: list[str] | None = None,
) -> pd.DataFrame:
    parts: list[pd.DataFrame] = []
    for path, config in workbook_entries:
        path = Path(path)
        if not path.exists():
            continue
        parts.append(import_valuation_workbook(path, config, sports=sports))

    if not parts:
        return pd.DataFrame()

    games = pd.concat(parts, ignore_index=True)
    return games.drop_duplicates(
        subset=["sport", "game_date", "home_team", "away_team", "network", "location_type"],
        keep="first",
    ).reset_index(drop=True)


def build_reference_tables(
    games: pd.DataFrame,
    workbook_paths: list[Path | str] | None = None,
    sports: list[str] | None = None,
    root: Path | str | None = None,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Derive full D1 team table and network reach from KEY tabs."""
    sport_list = sports or sorted(games["sport"].unique())
    root_path = Path(root or Path.cwd())
    teams = build_d1_teams(sport_list, games=games, root=root_path)

    key_df = merge_key_tabs(workbook_paths or [])
    if not key_df.empty:
        networks = build_networks_from_key(key_df, sport_list, games)
    else:
        network_rows: list[dict] = []
        for sport in sport_list:
            sport_games = games[games["sport"] == sport]
            network_rows.extend(compute_network_reach_scores(sport_games, sport))
        networks = pd.DataFrame(network_rows).drop_duplicates(subset=["network", "sport"])

    return teams, networks


def save_merged_import(
    workbook_entries: list[tuple[Path | str, ValuationImportConfig]],
    games_output: Path | str,
    teams_output: Path | str,
    networks_output: Path | str,
    sports: list[str] | None = None,
    root: Path | str | None = None,
) -> pd.DataFrame:
    games = import_valuation_workbooks(workbook_entries, sports=sports)
    workbook_paths = [path for path, _ in workbook_entries]
    root_path = Path(root or Path(games_output).resolve().parents[1])
    teams, networks = build_reference_tables(games, workbook_paths, sports, root=root_path)

    games_output = Path(games_output)
    teams_output = Path(teams_output)
    networks_output = Path(networks_output)
    games_output.parent.mkdir(parents=True, exist_ok=True)

    games.to_csv(games_output, index=False)
    teams.to_csv(teams_output, index=False)
    networks.to_csv(networks_output, index=False)
    return games


def load_workbook_entries_from_config(config: dict, root: Path | str) -> list[tuple[Path, ValuationImportConfig]]:
    root = Path(root)
    paths = config.get("paths", {})
    entries = paths.get("valuation_workbooks")
    if not entries:
        legacy = paths.get("arizona_workbook")
        if legacy:
            entries = [e for e in DEFAULT_VALUATION_WORKBOOKS if e["file"].endswith(Path(legacy).name)]
        else:
            entries = DEFAULT_VALUATION_WORKBOOKS

    result: list[tuple[Path, ValuationImportConfig]] = []
    for entry in entries:
        cfg = config_from_dict(entry)
        file_paths = entry.get("files")
        if file_paths:
            for file_path in file_paths:
                path = root / file_path
                if path.exists():
                    result.append((path, cfg))
            continue
        path = root / entry["file"]
        if path.exists():
            result.append((path, cfg))
    return result
