from __future__ import annotations

import re
from datetime import date, datetime, time
from pathlib import Path

import openpyxl
import pandas as pd

from viewership_model.data.key_tab_import import canonicalize_network_name
from viewership_model.data.network_reach import compute_network_reach_scores

ARIZONA = "Arizona"
PRIMARY_SHEET = "Valuation (24-25 Data)"
FALLBACK_SHEET = "Valuation 2024-25 Reg Season"
YELLOW_RGB = {"FFFFFF00", "FFFF00"}

NETWORK_ALIASES = {
    "accn": "ACC Network",
    "acc network": "ACC Network",
    "espn+": "ESPN+",
    "espn2": "ESPN2",
    "espnu": "ESPNU",
    "cbssn": "CBSSN",
    "fs1": "FS1",
    "fox": "FOX",
    "flocollege": "FloCollege",
    "nec front row": "NEC Front Row",
    "astros.com": "Astros.com",
    "peacock": "Peacock",
    "tnt": "TNT",
    "cbs": "CBS",
    "espn": "ESPN",
}

SPORT_ALIASES = {
    "gynmastics": "gymnastics",
    "gymnastics": "gymnastics",
    "softball": "softball",
    "football": "football",
    "baseball": "baseball",
    "volleyball": "volleyball",
    "soccer": "soccer",
    "basketball": "basketball",
}


def _cell_is_yellow(cell) -> bool:
    fill = cell.fill
    if not fill or fill.fill_type != "solid" or not fill.fgColor:
        return False
    rgb = getattr(fill.fgColor, "rgb", None)
    if rgb in YELLOW_RGB:
        return True
    return str(rgb).upper().endswith("FFFF00")


def _normalize_network(value: str | None) -> str | None:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    text = str(value).strip()
    if not text:
        return None
    return canonicalize_network_name(text)


def _normalize_sport(gender: str | None, sport: str | None) -> str | None:
    if sport is None or (isinstance(sport, float) and pd.isna(sport)):
        return None
    sport_key = SPORT_ALIASES.get(str(sport).strip().lower())
    if not sport_key:
        return str(sport).strip().lower()
    gender_text = str(gender or "").strip().lower()
    if sport_key == "basketball":
        if gender_text.startswith("women") or gender_text == "w":
            return "womens_basketball"
        return "mens_basketball"
    return sport_key


def _coerce_date(game_date: date | datetime | str | None) -> date | None:
    if game_date is None:
        return None
    if isinstance(game_date, datetime):
        return game_date.date()
    if isinstance(game_date, date):
        return game_date
    text = str(game_date).strip()
    if not text:
        return None
    return pd.to_datetime(text).date()


def _parse_season(game_date: date | datetime | str | None) -> int:
    parsed = _coerce_date(game_date)
    if parsed is None:
        return datetime.now().year
    return parsed.year if parsed.month >= 8 else parsed.year - 1


def _parse_week(game_date: date | datetime | str | None) -> int:
    parsed = _coerce_date(game_date)
    if parsed is None:
        return 1
    season_start = date(_parse_season(parsed), 8, 1)
    if parsed < season_start:
        season_start = date(parsed.year - 1, 8, 1)
    return max(1, (parsed - season_start).days // 7 + 1)


def _coerce_time(game_time: time | datetime | str | None) -> time | None:
    if game_time is None:
        return None
    if isinstance(game_time, datetime):
        return game_time.time()
    if isinstance(game_time, time):
        return game_time
    text = str(game_time).strip()
    if not text:
        return None
    parsed = pd.to_datetime(text)
    return parsed.time() if isinstance(parsed, datetime) else parsed.to_pydatetime().time()


def _is_prime_time(game_time: time | datetime | str | None) -> bool:
    parsed = _coerce_time(game_time)
    if parsed is None:
        return False
    return parsed.hour >= 18


def _location_type(home_or_away: str | None) -> str:
    text = str(home_or_away or "").strip().lower()
    if text == "away":
        return "away"
    if text == "neutral":
        return "neutral"
    return "home"


def _teams_from_row(opponent: str, home_or_away: str | None) -> tuple[str, str]:
    opponent = str(opponent).strip()
    location = _location_type(home_or_away)
    if location == "away":
        return opponent, ARIZONA
    return ARIZONA, opponent


def _header_map(ws) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        value = ws.cell(1, col).value
        if value is None:
            continue
        text = str(value).strip()
        mapping[text] = col
        if text != str(value):
            mapping[str(value)] = col
    return mapping


def _header_col(headers: dict[str, int], *names: str) -> int | None:
    for name in names:
        if name in headers:
            return headers[name]
        stripped = name.strip()
        if stripped in headers:
            return headers[stripped]
    lowered = {key.strip().lower(): idx for key, idx in headers.items()}
    for name in names:
        idx = lowered.get(name.strip().lower())
        if idx is not None:
            return idx
    return None


def _numeric_value(value) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        if value.startswith("="):
            return None
        try:
            return float(value.replace(",", ""))
        except ValueError:
            return None
    return None


def _read_sheet(ws, ws_values, source_sheet: str) -> list[dict]:
    headers = _header_map(ws)
    sport_col = _header_col(headers, "Sport")
    away_col = _header_col(headers, "Away ", "Away")
    network_col = _header_col(headers, "Network")
    viewers_col = _header_col(headers, "Avg Viewers")
    if None in (sport_col, away_col, network_col, viewers_col):
        missing = []
        if sport_col is None:
            missing.append("Sport")
        if away_col is None:
            missing.append("Away")
        if network_col is None:
            missing.append("Network")
        if viewers_col is None:
            missing.append("Avg Viewers")
        raise ValueError(f"Sheet '{source_sheet}' missing columns: {missing}")

    gender_col = _header_col(headers, "Gender")
    home_away_col = _header_col(headers, "Home or Away")
    date_col = _header_col(headers, "Date")
    time_col = _header_col(headers, "Time")
    conf_col = _header_col(headers, "Conf")
    conf_type_col = _header_col(headers, "Conference/Non Conference")
    arizona_rank_col = _header_col(headers, "Arizona Ranking")
    away_rank_col = _header_col(headers, "Away Rank")

    rows: list[dict] = []
    for row_idx in range(2, ws.max_row + 1):
        sport_raw = ws.cell(row_idx, sport_col).value
        if sport_raw is None or str(sport_raw).strip() == "":
            continue

        opponent = ws_values.cell(row_idx, away_col).value
        network = _normalize_network(ws_values.cell(row_idx, network_col).value)
        viewers_cell = ws.cell(row_idx, viewers_col)
        viewers = _numeric_value(ws_values.cell(row_idx, viewers_col).value)
        if opponent is None or network is None or viewers is None:
            continue

        gender = ws_values.cell(row_idx, gender_col).value if gender_col else None
        sport = _normalize_sport(gender, sport_raw)
        home_or_away = ws_values.cell(row_idx, home_away_col).value if home_away_col else "Home"
        game_date = ws_values.cell(row_idx, date_col).value if date_col else None
        game_time = ws_values.cell(row_idx, time_col).value if time_col else None
        conf = ws_values.cell(row_idx, conf_col).value if conf_col else None
        conf_type = ws_values.cell(row_idx, conf_type_col).value if conf_type_col else None
        arizona_rank = ws_values.cell(row_idx, arizona_rank_col).value if arizona_rank_col else None
        away_rank = ws_values.cell(row_idx, away_rank_col).value if away_rank_col else None

        home_team, away_team = _teams_from_row(str(opponent), home_or_away)
        location_type = _location_type(home_or_away)
        conference = str(conf).strip() if conf not in (None, "") else "Big 12"
        if conf_type and str(conf_type).strip().lower().startswith("non"):
            conference = "Non-Conference"

        season = _parse_season(game_date)
        week = _parse_week(game_date)
        is_ranked = int(
            pd.notna(arizona_rank) and str(arizona_rank).strip() != ""
            or pd.notna(away_rank) and str(away_rank).strip() != ""
        )

        rows.append(
            {
                "game_id": f"AZ-{sport}-{season}-{row_idx:04d}",
                "sport": sport,
                "season": season,
                "week": week,
                "home_team": home_team,
                "away_team": away_team,
                "network": network,
                "conference": conference,
                "location_type": location_type,
                "location_city": "Tucson" if home_team == ARIZONA else "Unknown",
                "location_state": "AZ" if home_team == ARIZONA else "ST",
                "is_rivalry": 0,
                "is_ranked_matchup": is_ranked,
                "is_prime_time": int(_is_prime_time(game_time)),
                "viewership_millions": round(viewers / 1_000_000, 6),
                "avg_viewers": viewers,
                "is_estimate": int(_cell_is_yellow(viewers_cell)),
                "game_date": _coerce_date(game_date),
                "gender": str(gender or "").strip(),
                "opponent": str(opponent).strip(),
                "source_sheet": source_sheet,
            }
        )
    return rows


def import_arizona_workbook(
    workbook_path: Path | str,
    sheet_name: str | None = None,
) -> pd.DataFrame:
    path = Path(workbook_path)
    wb = openpyxl.load_workbook(path, data_only=False)
    wb_values = openpyxl.load_workbook(path, data_only=True)

    sheets = [sheet_name] if sheet_name else [PRIMARY_SHEET]
    all_rows: list[dict] = []
    for sheet in sheets:
        if sheet not in wb.sheetnames:
            continue
        all_rows.extend(_read_sheet(wb[sheet], wb_values[sheet], sheet))

    df = pd.DataFrame(all_rows)
    if df.empty:
        return df

    df = df.drop_duplicates(
        subset=["sport", "game_date", "opponent", "network", "location_type"],
        keep="first",
    ).reset_index(drop=True)
    return df


def build_reference_tables(games: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Derive team and network reach tables from Arizona historical viewership."""
    teams_rows: list[dict] = []
    for sport in sorted(games["sport"].unique()):
        sport_games = games[games["sport"] == sport]
        arizona_avg = sport_games["avg_viewers"].mean()
        teams_rows.append(
            {
                "team": ARIZONA,
                "conference": "Big 12",
                "popularity_score": 85,
                "market_size_millions": 1.0,
                "sport": sport,
            }
        )

        opponents = sorted(set(sport_games["opponent"].unique()))
        for opponent in opponents:
            opp_games = sport_games[sport_games["opponent"] == opponent]
            avg_viewers = opp_games["avg_viewers"].mean()
            popularity = min(95, max(35, 20 * (avg_viewers / max(arizona_avg, 1)) ** 0.35))
            conf = opp_games["conference"].mode().iloc[0] if not opp_games.empty else "Unknown"
            teams_rows.append(
                {
                    "team": opponent,
                    "conference": conf,
                    "popularity_score": round(popularity, 1),
                    "market_size_millions": 1.0,
                    "sport": sport,
                }
            )

    teams = pd.DataFrame(teams_rows).drop_duplicates(subset=["team", "sport"])

    network_rows: list[dict] = []
    for sport in sorted(games["sport"].unique()):
        sport_games = games[games["sport"] == sport]
        network_rows.extend(compute_network_reach_scores(sport_games, sport))

    networks = pd.DataFrame(network_rows).drop_duplicates(subset=["network", "sport"])
    return teams, networks


def save_import(
    workbook_path: Path | str,
    games_output: Path | str,
    teams_output: Path | str,
    networks_output: Path | str,
) -> pd.DataFrame:
    games = import_arizona_workbook(workbook_path)
    teams, networks = build_reference_tables(games)

    games_output = Path(games_output)
    teams_output = Path(teams_output)
    networks_output = Path(networks_output)

    games_output.parent.mkdir(parents=True, exist_ok=True)
    games.to_csv(games_output, index=False)
    teams.to_csv(teams_output, index=False)
    networks.to_csv(networks_output, index=False)
    return games
