from __future__ import annotations

import math
import re
from pathlib import Path

import openpyxl
import pandas as pd

KEY_SHEET = "KEY"
HOUSEHOLD_CEILING = 88_000_000.0

# Map KEY tab labels and common game-sheet spellings to canonical network names.
CANONICAL_NETWORK_ALIASES: dict[str, str] = {
    "accn": "ACC Network",
    "acc network": "ACC Network",
    "accn+": "ACCN+",
    "astros.com": "Astros.com",
    "cbs": "CBS",
    "cbs (regular season)": "CBS",
    "cbs (post season)": "CBS",
    "cbssn": "CBSSN",
    "espn": "ESPN",
    "espn+": "ESPN+",
    "espn2": "ESPN2",
    "espnu": "ESPNU",
    "flocollege": "FloCollege",
    "flo college": "FloCollege",
    "fox": "FOX",
    "fox (regular season)": "FOX",
    "fox (post season)": "FOX",
    "fs1": "FS1",
    "fs2": "FS2",
    "btn": "BTN",
    "big ten network": "BTN",
    "big10+": "BTN",
    "bt n": "BTN",
    "hornnet sports network": "Hornet Sports Network",
    "hornet sports network": "Hornet Sports Network",
    "mountain west network": "Mountain West Network",
    "nec front row": "NEC Front Row",
    "pac-12 network": "Pac-12 Network",
    "pac 12 network": "Pac-12 Network",
    "peacock": "Peacock",
    "sec network": "SEC Network",
    "sec+": "SEC+",
    "sec network+": "SEC+",
    "tbs": "TBS",
    "tnt": "TNT",
    "trutv": "TruTV",
    "america east tv": "America East TV",
    "youtube": "YouTube",
}

DEFAULT_HOUSEHOLDS_BY_DISTRIBUTION: dict[str, float] = {
    "national linear": 35_000_000.0,
    "streaming": 1_500_000.0,
    "regional sports network (rsn)": 2_000_000.0,
    "regional sports network": 2_000_000.0,
}


def canonicalize_network_name(name: str) -> str:
    text = str(name).strip()
    if not text:
        return text
    lowered = text.lower()
    if lowered in CANONICAL_NETWORK_ALIASES:
        return CANONICAL_NETWORK_ALIASES[lowered]
    # Drop postseason / regular-season parentheticals for lookup.
    base = re.sub(r"\s*\((regular season|post season)\)\s*$", "", lowered, flags=re.I).strip()
    if base in CANONICAL_NETWORK_ALIASES:
        return CANONICAL_NETWORK_ALIASES[base]
    return text


def _distribution_key(distribution: str | None) -> str:
    if not distribution:
        return "streaming"
    return str(distribution).strip().lower()


def _estimate_households(
    households: float | None,
    distribution: str | None,
    cpm: float | None,
    network: str | None = None,
) -> float:
    if households is not None and not (isinstance(households, float) and math.isnan(households)):
        return float(households)
    if network == "BTN":
        return 40_000_000.0
    if network == "SEC Network":
        return 45_000_000.0
    if network == "FS2":
        return 35_000_000.0
    dist = _distribution_key(distribution)
    base = DEFAULT_HOUSEHOLDS_BY_DISTRIBUTION.get(dist, 1_000_000.0)
    if cpm is not None and not (isinstance(cpm, float) and math.isnan(cpm)):
        return base * (1.0 + float(cpm) / 80.0)
    return base


def reach_score_from_key_row(
    households: float | None,
    distribution: str | None,
    visibility: float | None,
    cpm: float | None,
    network: str | None = None,
) -> float:
    """Convert KEY tab network fields to a 0–100 reach score for the viewership formula."""
    hh = _estimate_households(households, distribution, cpm, network)
    vis = 0.4 if visibility is None or (isinstance(visibility, float) and math.isnan(visibility)) else float(visibility)
    cpm_val = 20.0 if cpm is None or (isinstance(cpm, float) and math.isnan(cpm)) else float(cpm)

    dist = _distribution_key(distribution)
    if network in {"BTN", "SEC Network", "FS2"}:
        dist = "national linear"
        if hh < 20_000_000:
            hh = {"BTN": 40_000_000.0, "SEC Network": 45_000_000.0, "FS2": 35_000_000.0}[network]

    if dist == "national linear":
        hh_ratio = min(1.0, hh / HOUSEHOLD_CEILING) ** 0.38
        viewership_factor = 0.30 + 0.55 * vis
        cpm_boost = 1.0 + 0.10 * (cpm_val / 40.0)
        reach = 100.0 * hh_ratio * viewership_factor * cpm_boost
    elif "rsn" in dist or "regional" in dist:
        hh_ratio = min(1.0, hh / HOUSEHOLD_CEILING) ** 0.35
        reach = 100.0 * hh_ratio * (0.22 + 0.35 * vis) * 0.85
    else:
        # Streaming: ESPN+ (~25M subs) anchors ~11; niche streams stay near 8.
        if hh >= 15_000_000:
            sub_ratio = (hh / 25_000_000) ** 0.25
            reach = 11.0 * sub_ratio * (0.85 + 0.35 * vis)
        elif hh >= 1_000_000:
            reach = 8.5 + 3.0 * (hh / 15_000_000) ** 0.35
        else:
            reach = 8.0 + 2.0 * min(1.0, hh / 1_000_000) ** 0.25

    return round(min(95.0, max(8.0, reach)), 1)


def parse_key_tab(workbook_path: Path | str) -> pd.DataFrame:
    """Read the KEY sheet network table from a valuation workbook."""
    path = Path(workbook_path)
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    if KEY_SHEET not in wb.sheetnames:
        wb.close()
        return pd.DataFrame()

    ws = wb[KEY_SHEET]
    rows: list[dict] = []
    in_networks = False
    for row in ws.iter_rows(values_only=True):
        row = tuple(row) if row else ()
        col0 = row[0] if len(row) > 0 else None
        if col0 == "Network":
            in_networks = True
            continue
        if not in_networks:
            continue
        if col0 == "Sport":
            break
        if not col0 or not str(col0).strip():
            continue

        network_name = str(col0).strip()
        households = row[1] if len(row) > 1 else None
        distribution = row[2] if len(row) > 2 else None
        visibility = row[3] if len(row) > 3 else None
        cpm = row[6] if len(row) > 6 else None

        canonical = canonicalize_network_name(network_name)
        reach = reach_score_from_key_row(households, distribution, visibility, cpm, canonical)
        rows.append(
            {
                "network": canonical,
                "key_name": network_name,
                "households": households,
                "distribution": distribution,
                "visibility": visibility,
                "cpm": cpm,
                "reach_score": reach,
                "source_workbook": path.name,
            }
        )
    wb.close()
    return pd.DataFrame(rows)


def merge_key_tabs(workbook_paths: list[Path | str]) -> pd.DataFrame:
    """Merge KEY tabs from multiple workbooks; prefer rows with subscriber data."""
    parts: list[pd.DataFrame] = []
    for path in workbook_paths:
        path = Path(path)
        if path.exists():
            part = parse_key_tab(path)
            if not part.empty:
                parts.append(part)
    if not parts:
        return pd.DataFrame()

    merged = pd.concat(parts, ignore_index=True)
    merged["_has_hh"] = merged["households"].notna().astype(int)
    merged = merged.sort_values(["network", "_has_hh", "reach_score"], ascending=[True, False, False])
    return merged.drop_duplicates(subset=["network"], keep="first").drop(columns=["_has_hh"]).reset_index(drop=True)


def build_networks_from_key(
    key_df: pd.DataFrame,
    sports: list[str],
    games: pd.DataFrame | None = None,
) -> pd.DataFrame:
    """Build per-sport networks.csv rows from KEY tab data, filling gaps from game history."""
    from viewership_model.data.network_reach import compute_network_reach_scores

    if key_df.empty and (games is None or games.empty):
        return pd.DataFrame(columns=["network", "reach_score", "sport"])

    key_networks = set(key_df["network"]) if not key_df.empty else set()
    rows: list[dict] = []

    for sport in sports:
        if not key_df.empty:
            for _, row in key_df.iterrows():
                rows.append(
                    {
                        "network": row["network"],
                        "reach_score": float(row["reach_score"]),
                        "sport": sport,
                    }
                )

        if games is not None and not games.empty:
            sport_games = games[games["sport"] == sport]
            missing = set(sport_games["network"].unique()) - key_networks
            if missing:
                supplemental = compute_network_reach_scores(
                    sport_games[sport_games["network"].isin(missing)],
                    sport,
                )
                rows.extend(supplemental)

    if not rows:
        return pd.DataFrame(columns=["network", "reach_score", "sport"])

    networks = pd.DataFrame(rows)
    return networks.drop_duplicates(subset=["network", "sport"], keep="first").reset_index(drop=True)
