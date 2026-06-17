from __future__ import annotations

import re
from pathlib import Path

import openpyxl
import pandas as pd

from viewership_model.data.arizona_import import (
    _coerce_date,
    _location_type,
    _numeric_value,
    _parse_season,
    _parse_week,
)
from viewership_model.data.d1_teams import normalize_team_name
from viewership_model.data.key_tab_import import canonicalize_network_name

MONMOUTH_SCHEDULE_SHEETS = [
    "MBB Schedule 2024-2025",
    "MBB Schedule 2025-26",
    "WBB Schedule 2024-25",
    "WBB Schedule 2025-26",
    "Baseball Schedule 2024-2025",
    "Football Schedule 2024",
    "Football Schedule 2025",
]

SHEET_SPORT = {
    "MBB": "mens_basketball",
    "WBB": "womens_basketball",
    "Football": "football",
    "Baseball": "baseball",
}


def _sport_from_sheet(sheet_name: str) -> str | None:
    for prefix, sport in SHEET_SPORT.items():
        if sheet_name.startswith(prefix):
            return sport
    return None


def _gender_from_sport(sport: str) -> str:
    if sport == "womens_basketball":
        return "Womens"
    if sport == "mens_basketball":
        return "Mens"
    return ""


def _clean_opponent(raw: str) -> str | None:
    text = str(raw).strip()
    if not text:
        return None
    if "exhibition" in text.lower():
        return None
    text = re.sub(r"\s*\*+\s*", " ", text)
    text = re.sub(r"\s*\([^)]*\)\s*", " ", text).strip()
    text = re.sub(r"\s+University$", "", text, flags=re.I)
    text = re.sub(r"\s+", " ", text).strip()
    return normalize_team_name(text) if text else None


def _normalize_monmouth_network(value: str | None) -> str | None:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    text = str(value).strip()
    if not text:
        return None
    lowered = text.lower()
    aliases = {
        "big10+": "BTN",
        "big ten network": "BTN",
        "big10 network": "BTN",
        "flosports": "FloSports",
        "flo sports": "FloSports",
        "sec network": "SEC Network",
        "nbc sports philly": "NBC Sports Philly",
        "sny": "SNY",
        "jerseyjam.net": "JerseyJam.net",
        "usa baseball.tv": "USA Baseball.tv",
    }
    if lowered in aliases:
        return aliases[lowered]
    return canonicalize_network_name(text)


def _broadcast_rows(row: tuple) -> list[tuple[str, float]]:
    pairs: list[tuple[str | None, object]] = [
        (row[5] if len(row) > 5 else None, row[6] if len(row) > 6 else None),
        (row[7] if len(row) > 7 else None, row[8] if len(row) > 8 else None),
        (row[9] if len(row) > 9 else None, row[10] if len(row) > 10 else None),
    ]
    out: list[tuple[str, float]] = []
    for network_raw, viewers_raw in pairs:
        network = _normalize_monmouth_network(network_raw)
        viewers = _numeric_value(viewers_raw)
        if network and viewers is not None and viewers > 0:
            out.append((network, float(viewers)))
    return out


def _teams_for_game(
    home_team: str,
    opponent: str,
    home_away_neutral: str | None,
) -> tuple[str, str, str]:
    location = str(home_away_neutral or "Home").strip()
    loc_type = _location_type(location)
    if loc_type == "away":
        return opponent, home_team, loc_type
    if loc_type == "neutral":
        return home_team, opponent, loc_type
    return home_team, opponent, loc_type


def import_monmouth_schedule_workbook(
    workbook_path: Path | str,
    home_team: str = "Monmouth",
    sheets: list[str] | None = None,
    game_id_prefix: str = "MU",
    default_conference: str = "CAA",
    home_city: str = "West Long Branch",
    home_state: str = "NJ",
) -> pd.DataFrame:
    path = Path(workbook_path)
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    sheet_names = sheets or [
        sn for sn in wb.sheetnames if _sport_from_sheet(sn) is not None
    ]

    rows: list[dict] = []
    for sheet_name in sheet_names:
        if sheet_name not in wb.sheetnames:
            continue
        sport = _sport_from_sheet(sheet_name)
        if not sport:
            continue
        ws = wb[sheet_name]
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or len(row) < 7:
                continue
            opponent = _clean_opponent(row[1] if len(row) > 1 else "")
            if not opponent:
                continue
            game_date = _coerce_date(row[0] if len(row) > 0 else None)
            broadcasts = _broadcast_rows(row)
            if not broadcasts:
                continue

            home, away, loc_type = _teams_for_game(
                home_team, opponent, row[2] if len(row) > 2 else "Home"
            )
            season = _parse_season(game_date)
            week = _parse_week(game_date)
            gender = _gender_from_sport(sport)

            for network, viewers in broadcasts:
                rows.append(
                    {
                        "game_id": f"{game_id_prefix}-{sport}-{season}-{row_idx:04d}-{network[:3]}",
                        "sport": sport,
                        "season": season,
                        "week": week,
                        "home_team": home,
                        "away_team": away,
                        "network": network,
                        "conference": default_conference,
                        "location_type": loc_type,
                        "location_city": home_city if home == home_team else "Unknown",
                        "location_state": home_state if home == home_team else "ST",
                        "is_rivalry": 0,
                        "is_ranked_matchup": 0,
                        "is_prime_time": 0,
                        "viewership_millions": round(viewers / 1_000_000, 6),
                        "avg_viewers": viewers,
                        "is_estimate": 0,
                        "game_date": game_date,
                        "gender": gender,
                        "opponent": opponent,
                        "source_sheet": f"{path.name}:{sheet_name}",
                    }
                )

    wb.close()
    if not rows:
        return pd.DataFrame()

    df = pd.DataFrame(rows)
    return df.drop_duplicates(
        subset=["sport", "game_date", "home_team", "away_team", "network", "location_type"],
        keep="first",
    ).reset_index(drop=True)
